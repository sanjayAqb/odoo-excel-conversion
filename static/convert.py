"""
Convert supplier product sheets into the Odoo import template format.
Runs in the browser via Pyodide.
"""

from __future__ import annotations

import csv
import io
import json
import re
from typing import Any

from openpyxl import load_workbook

from xlsx_styles import (
    CURRENCY_FORMAT,
    PERCENT_FORMAT,
    autofit_columns,
    freeze_below_header,
    style_data_cell,
    style_header_cell,
)

REQUIRED_SUPPLIER_KEYS = [
    "name",
    "barcode",
    "cost/case",
    "selling price",
    "plu",
    "markup",
    "vat",
]

ODOO_CATEGORY_REQUIRED = {
    "parent category/category name": "Parent Category/Category Name",
    "category name": "Category Name",
    "external id": "External ID",
}


def _clean_name(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _normalize_header(value: Any) -> str:
    text = _clean_name(value).lower()
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _looks_like_odoo_category_header(cells: list[Any]) -> bool:
    normalized = {_normalize_header(c) for c in cells if _normalize_header(c)}
    return "category name" in normalized and "external id" in normalized


def _resolve_odoo_category_columns(headers: list[str]) -> dict[str, str]:
    by_norm = {_normalize_header(h): h for h in headers if _normalize_header(h)}
    resolved: dict[str, str] = {}
    for norm_key, label in ODOO_CATEGORY_REQUIRED.items():
        if norm_key in by_norm:
            resolved[norm_key] = by_norm[norm_key]
    return resolved


def detect_odoo_category_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:20]):
        if _looks_like_odoo_category_header(row):
            return idx
    raise ValueError(
        "Could not find a header row in the Odoo category export. "
        "The file must contain: Parent Category/Category Name, Category Name, External ID."
    )


def parse_odoo_category_export(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        raise ValueError("The Odoo category export file is empty.")

    header_idx = detect_odoo_category_header_row(rows)
    raw_headers = rows[header_idx]
    headers: list[str] = []
    for i, h in enumerate(raw_headers):
        label = str(h).strip() if h is not None else ""
        headers.append(label or f"Column_{i + 1}")

    resolved = _resolve_odoo_category_columns(headers)
    missing = [label for key, label in ODOO_CATEGORY_REQUIRED.items() if key not in resolved]
    if missing:
        raise ValueError(
            "The uploaded category file is missing required column(s): "
            + ", ".join(missing)
            + ". The file must contain: Parent Category/Category Name, Category Name, External ID."
        )

    parent_col = resolved["parent category/category name"]
    name_col = resolved["category name"]
    external_col = resolved["external id"]

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        record: dict[str, Any] = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            record[header] = "" if val is None else val

        if not _clean_name(record.get(name_col, "")):
            continue
        if not _clean_name(record.get(external_col, "")):
            continue

        records.append(
            {
                parent_col: _clean_name(record.get(parent_col, "")),
                name_col: _clean_name(record.get(name_col, "")),
                external_col: _clean_name(record.get(external_col, "")),
            }
        )

    if not records:
        raise ValueError("No category rows found in the Odoo category export file.")

    return headers, records


def build_category_lookups(
    headers: list[str],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    resolved = _resolve_odoo_category_columns(headers)
    parent_col = resolved["parent category/category name"]
    name_col = resolved["category name"]
    external_col = resolved["external id"]

    parent_lookup: dict[str, str] = {}
    child_lookup: dict[tuple[str, str], str] = {}

    for record in records:
        parent_name = _clean_name(record.get(parent_col, ""))
        category_name = _clean_name(record.get(name_col, ""))
        external_id = _clean_name(record.get(external_col, ""))
        if not category_name or not external_id:
            continue
        if parent_name:
            child_lookup[(_normalize(parent_name), _normalize(category_name))] = external_id
        else:
            parent_lookup[_normalize(category_name)] = external_id

    return {
        "parent_lookup": parent_lookup,
        "child_lookup": child_lookup,
    }


def resolve_pos_category_external_id(
    department: str,
    sub_department: str,
    lookups: dict[str, Any],
) -> str:
    dept = _clean_name(department)
    sub = _clean_name(sub_department)
    parent_lookup = lookups.get("parent_lookup", {})
    child_lookup = lookups.get("child_lookup", {})

    if sub:
        return child_lookup.get((_normalize(dept), _normalize(sub)), "")

    if dept:
        return parent_lookup.get(_normalize(dept), "")

    return ""


def _as_text_identifier(value: Any) -> str:
    """Preserve barcodes/PLUs as full digit strings (never scientific notation)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    text = str(value).strip()
    if not text:
        return ""
    upper = text.upper()
    if "E+" in upper or "E-" in upper:
        try:
            as_float = float(text)
            if as_float == int(as_float):
                return str(int(as_float))
        except (ValueError, OverflowError):
            pass
    if re.fullmatch(r"-?\d+\.0+", text):
        return text.split(".")[0]
    return text


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _looks_like_header(cells: list[Any]) -> bool:
    normalized = {_normalize(c) for c in cells if _normalize(c)}
    return "name" in normalized and (
        "barcode" in normalized or "plu" in normalized or "selling price" in normalized
    )


def _resolve_required_aliases(headers: list[str]) -> dict[str, str]:
    """Map logical required keys to actual header strings found in the file."""
    found: dict[str, str] = {}
    for header in headers:
        key = _normalize(header)
        if not key:
            continue
        if key == "name":
            found["name"] = header
        elif key == "barcode":
            found["barcode"] = header
        elif key in ("cost/case", "cost case"):
            found["cost/case"] = header
        elif key == "selling price":
            found["selling price"] = header
        elif key == "plu":
            found["plu"] = header
        elif key in ("markup", "markup(%)", "markup %"):
            found["markup"] = header
        elif key in ("vat (%)", "vat(%)", "vat %", "vat"):
            found["vat"] = header
    return found


def detect_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:50]):
        if _looks_like_header(row):
            return idx
    raise ValueError(
        "Could not find a header row containing product columns "
        "(expected at least Name plus Barcode, PLU, or Selling Price)."
    )


def parse_supplier_rows(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        raise ValueError("The uploaded file is empty.")

    header_idx = detect_header_row(rows)
    raw_headers = rows[header_idx]
    headers: list[str] = []
    for i, h in enumerate(raw_headers):
        label = str(h).strip() if h is not None else ""
        headers.append(label or f"Column_{i + 1}")

    resolved = _resolve_required_aliases(headers)
    missing = [k for k in REQUIRED_SUPPLIER_KEYS if k not in resolved]
    if missing:
        label_map = {
            "name": "Name",
            "barcode": "Barcode",
            "cost/case": "Cost/Case",
            "selling price": "Selling Price",
            "plu": "PLU",
            "markup": "Markup / Markup(%)",
            "vat": "VAT (%) / VAT(%)",
        }
        pretty = ", ".join(label_map[m] for m in missing)
        raise ValueError(
            f"Missing required supplier column(s): {pretty}. "
            f"Found headers: {', '.join(h for h in headers if h and not h.startswith('Column_'))}"
        )

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # Skip footer / non-product lines
        first = str(row[0]).strip().lower() if row and row[0] is not None else ""
        if first.startswith("page ") or first == "products":
            continue

        record: dict[str, Any] = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val is None:
                record[header] = ""
            elif isinstance(val, float) and val == int(val):
                record[header] = int(val)
            else:
                record[header] = val
            key = _normalize(header)
            if key in ("barcode", "plu"):
                record[header] = _as_text_identifier(record[header])

        name_val = record.get(resolved["name"], "")
        if str(name_val).strip() == "":
            continue
        records.append(record)

    if not records:
        raise ValueError("No product rows found under the detected header row.")

    return headers, records


def _pick_source_header(headers: list[str], preferred_keys: list[str]) -> str | None:
    """
    Choose the best matching supplier header for a target Odoo field.
    preferred_keys are normalized names ordered from most to least specific
    (e.g. prefer 'vat (%)' over a bare 'vat' label column).
    """
    by_norm = {_normalize(h): h for h in headers if _normalize(h)}
    for key in preferred_keys:
        if key in by_norm:
            return by_norm[key]
    return None


def map_to_odoo(
    headers: list[str],
    records: list[dict[str, Any]],
    category_lookups: dict[str, Any] | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    odoo_headers = [
        "Name",
        "Internal Reference",
        "Cost",
        "Sales Price",
        "Point of Sale Category / External ID",
        "Track Inventory",
        "Product Category",
        "Barcode",
        "VAT %",
        "Markup %",
        "Brand Name",
        "Gross Margin %",
        "U Cost Price",
        "Exclude from API Sync",
    ]

    # Prefer specific % / price columns when ambiguous headers exist (e.g. VAT(%) vs VAT).
    field_preferences: dict[str, list[str]] = {
        "Name": ["name"],
        "Barcode": ["barcode"],
        "Cost": ["cost/case", "cost case", "cost"],
        "Sales Price": ["selling price"],
        "Internal Reference": ["plu"],
        "Markup %": ["markup(%)", "markup %", "markup"],
        "VAT %": ["vat (%)", "vat(%)", "vat %", "vat"],
    }

    source_for_odoo: dict[str, str] = {}
    for odoo_field, keys in field_preferences.items():
        chosen = _pick_source_header(headers, keys)
        if chosen:
            source_for_odoo[odoo_field] = chosen

    dept_col = _pick_source_header(headers, ["department"])
    sub_col = _pick_source_header(headers, ["sub-department"])

    mapped: list[dict[str, Any]] = []
    for record in records:
        out: dict[str, Any] = {h: "" for h in odoo_headers}
        for dst, src in source_for_odoo.items():
            val = record.get(src, "")
            if val is None:
                val = ""
            if dst in ("Barcode", "Internal Reference"):
                val = _as_text_identifier(val)
            out[dst] = val
        # Static Odoo defaults (not sourced from supplier sheet)
        out["Track Inventory"] = 1
        out["Product Category"] = "All"
        out["Exclude from API Sync"] = "TRUE"
        if category_lookups:
            dept = record.get(dept_col, "") if dept_col else ""
            sub = record.get(sub_col, "") if sub_col else ""
            out["Point of Sale Category / External ID"] = resolve_pos_category_external_id(
                dept, sub, category_lookups
            )
        mapped.append(out)
    return odoo_headers, mapped


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        if value.strip() == "":
            return None
        # Write numeric types for pure numeric strings (keep original text otherwise)
        text = value.strip()
        if re.fullmatch(r"-?\d+", text):
            return int(text)
        if re.fullmatch(r"-?\d+\.\d+", text):
            return float(text)
        return value
    return value


def _emit_progress(pct: int, progress_cb: Any = None) -> None:
    value = int(pct)
    if progress_cb is not None:
        progress_cb(value)
        return
    try:
        from js import window  # type: ignore

        window.__setConversionProgress(value)
    except Exception:
        pass


def build_xlsx_bytes(
    template_bytes: bytes,
    mapped_rows: list[dict[str, Any]],
    progress_cb: Any = None,
    progress_start: int = 40,
    progress_end: int = 85,
) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    # Find header row (contains "Name")
    header_row_idx = None
    header_to_col: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        values = [c.value for c in row]
        if any(_normalize(v) == "name" for v in values):
            header_row_idx = row[0].row
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    header_to_col[str(cell.value).strip()] = cell.column
            break

    if header_row_idx is None:
        raise ValueError("Odoo template is missing a header row with a Name column.")

    # Apply consistent header styling across all template columns
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row_idx, column=col)
        if cell.value is None or str(cell.value).strip() == "":
            continue
        style_header_cell(cell)

    # Delete existing data rows below the header, preserving column structure/styles on header
    if ws.max_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, ws.max_row - header_row_idx)

    mapped_fields = {
        "Name",
        "Barcode",
        "Cost",
        "Sales Price",
        "Internal Reference",
        "Markup %",
        "VAT %",
        "Track Inventory",
        "Product Category",
        "Exclude from API Sync",
        "Point of Sale Category / External ID",
    }

    TEXT_FIELDS = {"Barcode", "Internal Reference"}
    CURRENCY_FIELDS = {"Cost", "Sales Price", "U Cost Price"}
    PERCENT_FIELDS = {"Markup %", "VAT %", "Gross Margin %"}

    total = max(len(mapped_rows), 1)
    for i, record in enumerate(mapped_rows):
        excel_row = header_row_idx + 1 + i
        zebra = i % 2 == 1
        for field in mapped_fields:
            col = header_to_col.get(field)
            if col is None:
                continue
            if field in TEXT_FIELDS:
                cell = ws.cell(row=excel_row, column=col)
                text = _as_text_identifier(record.get(field))
                cell.value = text or None
                cell.number_format = "@"
            else:
                cell = ws.cell(row=excel_row, column=col, value=_cell_value(record.get(field)))
                if field in CURRENCY_FIELDS:
                    cell.number_format = CURRENCY_FORMAT
                elif field in PERCENT_FIELDS:
                    cell.number_format = PERCENT_FORMAT
            style_data_cell(cell, zebra=zebra)

        if i == 0 or (i + 1) % 25 == 0 or i + 1 == total:
            pct = progress_start + int((i + 1) / total * (progress_end - progress_start))
            _emit_progress(pct, progress_cb)

    autofit_columns(ws, min_row=header_row_idx)
    freeze_below_header(ws, header_row_idx)

    buf = io.BytesIO()
    wb.save(buf)
    _emit_progress(progress_end, progress_cb)
    return buf.getvalue()


def build_csv_text(odoo_headers: list[str], mapped_rows: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=odoo_headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in mapped_rows:
        writer.writerow({h: "" if row.get(h) is None else row.get(h) for h in odoo_headers})
    return buf.getvalue()


def convert(
    supplier_rows_json: str,
    template_bytes: bytes,
    detected_format: str,
    progress_cb: Any = None,
) -> str:
    """
    Main entry point for the browser UI.
    supplier_rows_json: JSON list of rows (list of lists)
    Returns JSON string with preview + base64 xlsx + csv text.
    progress_cb: optional callable(int) with 0–100 progress.
                 If omitted, falls back to window.__setConversionProgress in browser.
    """
    import base64

    _emit_progress(5, progress_cb)
    rows = json.loads(supplier_rows_json)
    _emit_progress(12, progress_cb)
    headers, records = parse_supplier_rows(rows)
    _emit_progress(28, progress_cb)
    odoo_headers, mapped = map_to_odoo(headers, records)
    _emit_progress(40, progress_cb)
    xlsx_bytes = build_xlsx_bytes(template_bytes, mapped, progress_cb=progress_cb)
    _emit_progress(90, progress_cb)
    csv_text = build_csv_text(odoo_headers, mapped)
    _emit_progress(96, progress_cb)

    preview_limit = 100
    result = {
        "ok": True,
        "detected_format": detected_format,
        "supplier_headers": headers,
        "row_count": len(mapped),
        "odoo_headers": odoo_headers,
        "preview": mapped[:preview_limit],
        "xlsx_base64": base64.b64encode(xlsx_bytes).decode("ascii"),
        "csv_text": csv_text,
    }
    _emit_progress(100, progress_cb)
    return json.dumps(result)
