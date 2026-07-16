"""
Generate Odoo Product Category import files from supplier sheets.
Runs in the browser via Pyodide.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from openpyxl import Workbook

from xlsx_styles import autofit_columns, freeze_below_header, style_data_cell, style_header_cell

CATEGORY_HEADERS = [
    "External ID",
    "Category Name",
    "Parent Category / External ID",
]


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_name(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _looks_like_category_header(cells: list[Any]) -> bool:
    normalized = {_normalize(c) for c in cells if _normalize(c)}
    return "department" in normalized and "sub-department" in normalized


def detect_category_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:50]):
        if _looks_like_category_header(row):
            return idx
    raise ValueError(
        "Could not find a header row with Department and Sub-Department columns."
    )


def parse_category_source_rows(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        raise ValueError("The uploaded file is empty.")

    header_idx = detect_category_header_row(rows)
    raw_headers = rows[header_idx]
    headers: list[str] = []
    for i, h in enumerate(raw_headers):
        label = str(h).strip() if h is not None else ""
        headers.append(label or f"Column_{i + 1}")

    dept_col = None
    sub_col = None
    for header in headers:
        key = _normalize(header)
        if key == "department":
            dept_col = header
        elif key == "sub-department":
            sub_col = header

    if not dept_col or not sub_col:
        raise ValueError(
            "Missing required columns: Department and Sub-Department. "
            f"Found headers: {', '.join(h for h in headers if h and not h.startswith('Column_'))}"
        )

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        first = _normalize(row[0] if row else "")
        if first.startswith("page "):
            continue

        record: dict[str, Any] = {}
        for i, header in enumerate(headers):
            val = row[i] if i < len(row) else None
            record[header] = "" if val is None else val

        dept = _clean_name(record.get(dept_col, ""))
        sub = _clean_name(record.get(sub_col, ""))
        if not dept and not sub:
            continue

        records.append({dept_col: dept, sub_col: sub})

    if not records:
        raise ValueError("No category data found under Department / Sub-Department columns.")

    return headers, records


def slugify_category_name(name: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "_", name.strip().lower())
    text = text.strip("_")
    return text or "category"


def make_external_id(name: str, used: set[str], parent_name: str | None = None) -> str:
    """Generate Odoo external ID from category name: cat_{slug}."""
    base = f"cat_{slugify_category_name(name)}"
    if parent_name and base in used:
        base = f"cat_{slugify_category_name(parent_name)}_{slugify_category_name(name)}"
    candidate = base
    suffix = 2
    while candidate in used:
        candidate = f"{base}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def build_category_records(
    records: list[dict[str, Any]],
) -> tuple[list[str], list[dict[str, Any]]]:
    if not records:
        raise ValueError("No category records to process.")

    dept_key, sub_key = list(records[0].keys())

    departments: set[str] = set()
    child_pairs: set[tuple[str, str]] = set()

    for record in records:
        dept = _clean_name(record.get(dept_key, ""))
        sub = _clean_name(record.get(sub_key, ""))
        if dept:
            departments.add(dept)
        if dept and sub:
            child_pairs.add((dept, sub))

    if not departments and not child_pairs:
        raise ValueError("No valid Department or Sub-Department values found.")

    used_ids: set[str] = set()
    dept_external: dict[str, str] = {}
    mapped: list[dict[str, Any]] = []

    children_by_dept: dict[str, list[str]] = {}
    for dept, sub in child_pairs:
        children_by_dept.setdefault(dept, []).append(sub)

    for dept in sorted(departments):
        external_id = make_external_id(dept, used_ids)
        dept_external[dept] = external_id
        mapped.append(
            {
                "External ID": external_id,
                "Category Name": dept,
                "Parent Category / External ID": "",
            }
        )

        parent_external = dept_external[dept]
        for sub in sorted(children_by_dept.get(dept, [])):
            external_id = make_external_id(sub, used_ids, parent_name=dept)
            mapped.append(
                {
                    "External ID": external_id,
                    "Category Name": sub,
                    "Parent Category / External ID": parent_external,
                }
            )

    return CATEGORY_HEADERS, mapped


def _read_template_headers(template_bytes: bytes) -> list[str]:
    text = template_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    headers = next(reader, None)
    if not headers:
        raise ValueError("Category template is empty.")
    cleaned = [h.strip() for h in headers if h is not None and str(h).strip()]
    if cleaned != CATEGORY_HEADERS:
        # Preserve template order but ensure required columns exist
        for required in CATEGORY_HEADERS:
            if required not in cleaned:
                raise ValueError(f"Category template is missing required column: {required}")
    return CATEGORY_HEADERS


def build_category_csv_text(headers: list[str], mapped_rows: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in mapped_rows:
        writer.writerow({h: "" if row.get(h) is None else row.get(h) for h in headers})
    return buf.getvalue()


def build_category_xlsx_bytes(headers: list[str], mapped_rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Categories"

    for col, header in enumerate(headers, start=1):
        style_header_cell(ws.cell(row=1, column=col, value=header))

    for row_idx, record in enumerate(mapped_rows, start=2):
        zebra = (row_idx - 2) % 2 == 1
        for col, header in enumerate(headers, start=1):
            value = record.get(header, "")
            cell = ws.cell(row=row_idx, column=col, value="" if value is None else value)
            style_data_cell(cell, zebra=zebra)

    autofit_columns(ws)
    freeze_below_header(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
