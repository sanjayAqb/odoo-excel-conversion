"""Shared Excel styling for Odoo export files."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_THIN = Side(style="thin", color="B0B0B0")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D3D3D3")
HEADER_FONT = Font(bold=True, color="000000")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="center")
ZEBRA_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")

CURRENCY_FORMAT = "#,##0.00"
PERCENT_FORMAT = "0.00"


def style_header_cell(cell: Any) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGNMENT
    cell.border = CELL_BORDER


def style_data_cell(cell: Any, *, zebra: bool = False) -> None:
    cell.border = CELL_BORDER
    cell.alignment = DATA_ALIGNMENT
    if zebra:
        cell.fill = ZEBRA_FILL


def autofit_columns(
    ws: Worksheet,
    *,
    min_row: int = 1,
    max_row: int | None = None,
    min_width: float = 10,
    max_width: float = 48,
    padding: int = 2,
) -> None:
    end_row = max_row or ws.max_row
    for col_cells in ws.iter_cols(min_row=min_row, max_row=end_row):
        col_letter = get_column_letter(col_cells[0].column)
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(longest + padding, min_width), max_width)


def freeze_below_header(ws: Worksheet, header_row: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
